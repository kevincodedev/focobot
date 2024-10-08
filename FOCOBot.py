import os
from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import (
    Application,
    CommandHandler,
    CallbackQueryHandler,
    MessageHandler,
    filters,
)
import asyncio
import nest_asyncio
import openpyxl

# Aplicar nest_asyncio para evitar problemas con bucles de eventos en ciertos entornos.
nest_asyncio.apply()

# Imprimir la ruta actual de trabajo para depuración
print(f"Ruta actual de trabajo: {os.getcwd()}")

# Función para inicializar el archivo Excel con las 6 columnas necesarias.
def initialize_excel(file_path):
    if not os.path.exists(file_path):
        wb = openpyxl.Workbook()
        hoja = wb.active
        hoja.title = "DatosUsuarios"
        # Crear los encabezados
        hoja.append([
            "User ID",
            "Usuario de Telegram",
            "Pregunta 1 Seleccionada",
            "Pregunta 2 Seleccionada",
            "Pregunta 3 Seleccionada",
            "Pregunta 4 Seleccionada"
        ])
        wb.save(file_path)
        print(f"Archivo Excel creado en: {file_path}")  # Depuración

# Función para manejar el comando /start
async def start(update: Update, context):
    user = update.effective_user
    user_id = user.id
    username = user.username if user.username else user.first_name

    # Ruta del archivo Excel
    excel_path = os.path.join(os.getcwd(), "datos_usuarios.xlsx")
    initialize_excel(excel_path)

    # Cargar el libro de Excel
    try:
        wb = openpyxl.load_workbook(excel_path)
        hoja = wb["DatosUsuarios"]
    except Exception as e:
        await update.message.reply_text('Hubo un error al acceder a la base de datos.')
        print(f"Error al cargar el libro de Excel: {e}")  # Depuración
        return

    # Verificar si el usuario ya existe
    user_exists = False
    for row in hoja.iter_rows(min_row=2, values_only=True):
        if row[0] == user_id:
            user_exists = True
            break

    if not user_exists:
        # Si el usuario no existe, agregar una nueva fila con contadores en 0
        hoja.append([user_id, username, 0, 0, 0, 0])
        wb.save(excel_path)
        print(f"Nuevo usuario agregado: {username} con User ID: {user_id}")  # Depuración

    # Mensaje de bienvenida y menú de preguntas
    await update.message.reply_text('¡Bienvenido! Puedes seleccionar una pregunta usando el menú:')
    keyboard = [
        [InlineKeyboardButton("Obtener enlace al GIEP", callback_data='1')],
        [InlineKeyboardButton("Si estoy de vacaciones, ¿cuándo entrego mis tareas?", callback_data='2')],
        [InlineKeyboardButton("¿Cuándo se publican los nuevos e-books?", callback_data='3')],
        [InlineKeyboardButton("¿En qué parte del telegram se encuentran los e-books?", callback_data='4')],
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)
    await update.message.reply_text('Selecciona una pregunta usando el menú:', reply_markup=reply_markup)

# Función para manejar los botones del menú
async def button(update: Update, context):
    query = update.callback_query
    await query.answer()

    user = update.effective_user
    user_id = user.id
    username = user.username if user.username else user.first_name

    print(f"Usuario ID: {user_id}, Username: {username}")  # Depuración

    # Diccionario para mapear los callback_data a las preguntas
    questions = {
        '1': "Has seleccionado: Obtener enlace al GIEP",
        '2': "Has seleccionado: Si estoy de vacaciones, ¿cuándo entrego mis tareas?",
        '3': "Has seleccionado: ¿Cuándo se publican los nuevos e-books?",
        '4': "Has seleccionado: ¿En qué parte del telegram se encuentran los e-books?"
    }

    # Diccionario para mapear los callback_data a las respuestas
    answers = {
        '1': "Enlace al GIEP: https://www.giep-platform.pafar.com.ve",
        '2': "Las personas que se encuentran de vacaciones tienen una semana para entregar sus actividades, después de integrarse de manera física a Movilnet.",
        '3': "Los nuevos e-books se publican cada 2 semanas después de la entrega, es decir, 15 días o 10 días hábiles.",
        '4': "En los grupos de Telegram y en los correos corporativos (@movilnet.com.ve)."
    }

    selected_option = query.data

    question_selected = questions.get(selected_option, "Opción no válida")
    answer = answers.get(selected_option, "No hay respuesta para esta opción")

    # Enviar la respuesta al usuario
    await query.edit_message_text(text=f"<b>{question_selected}</b>\n{answer}", parse_mode="HTML")

    try:
        # Actualizar el contador en el Excel
        excel_path = os.path.join(os.getcwd(), "datos_usuarios.xlsx")
        wb = openpyxl.load_workbook(excel_path)
        hoja = wb["DatosUsuarios"]

        # Buscar la fila del usuario y actualizar el contador correspondiente
        user_found = False
        for row in hoja.iter_rows(min_row=2):
            print(f"Revisando fila con User ID: {row[0].value}")  # Depuración
            if row[0].value == user_id:
                user_found = True
                if selected_option == '1':
                    hoja.cell(row=row[0].row, column=3, value=(row[2].value or 0) + 1)
                    print(f"Incrementando Pregunta 1 a {row[2].value + 1}")  # Depuración
                elif selected_option == '2':
                    hoja.cell(row=row[0].row, column=4, value=(row[3].value or 0) + 1)
                    print(f"Incrementando Pregunta 2 a {row[3].value + 1}")  # Depuración
                elif selected_option == '3':
                    hoja.cell(row=row[0].row, column=5, value=(row[4].value or 0) + 1)
                    print(f"Incrementando Pregunta 3 a {row[4].value + 1}")  # Depuración
                elif selected_option == '4':
                    hoja.cell(row=row[0].row, column=6, value=(row[5].value or 0) + 1)
                    print(f"Incrementando Pregunta 4 a {row[5].value + 1}")  # Depuración
                break

        if not user_found:
            print("Usuario no encontrado en el archivo Excel.")  # Depuración

        wb.save(excel_path)
        print(f"Contadores actualizados para User ID: {user_id}")  # Depuración

    except Exception as e:
        await query.message.reply_text('Hubo un error al actualizar tus datos. Por favor, inténtalo de nuevo más tarde.')
        print(f"Error al actualizar el Excel: {e}")  # Depuración

    # Mostrar el menú de nuevo
    keyboard = [
        [InlineKeyboardButton("Obtener enlace al GIEP", callback_data='1')],
        [InlineKeyboardButton("Si estoy de vacaciones, ¿cuándo entrego mis tareas?", callback_data='2')],
        [InlineKeyboardButton("¿Cuándo se publican los nuevos e-books?", callback_data='3')],
        [InlineKeyboardButton("¿En qué parte del telegram se encuentran los e-books?", callback_data='4')],
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)
    await query.message.reply_text('Selecciona otra pregunta usando el menú:', reply_markup=reply_markup)

# Función para mostrar el menú cuando el usuario envía cualquier otro mensaje
async def show_menu_on_first_message(update: Update, context):
    await start(update, context)

# Función principal para configurar y ejecutar el bot
async def main():
    # Asignar el token directamente (Asegúrate de haber regenerado el token)
    TOKEN = "7047103350:AAFyzHfZcG0DVxq515V77Vrnd1AeroG5adE"

    if not TOKEN:
        print("Error: El token del bot no está definido. Asegúrate de establecer la variable de entorno TELEGRAM_BOT_TOKEN.")
        return

    # Crear la aplicación del bot utilizando el token directamente
    application = Application.builder().token(TOKEN).build()

    # Añadir manejadores
    application.add_handler(CommandHandler("start", start))
    application.add_handler(CallbackQueryHandler(button))
    application.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, show_menu_on_first_message))

    # Iniciar el bot con polling
    await application.run_polling()

# Punto de entrada del script
if __name__ == '__main__':
    asyncio.run(main())
